from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook("reports/barchart.xlsx")
sh = wb["Report"]

# Use direct references from `sh` for min and max rows/columns
min_column = sh.min_column
max_column = sh.max_column
min_row = sh.min_row
max_row = sh.max_row

# Adding sum formulas to each column in the next row after the last
for i in range(min_column + 1, max_column + 1):
    col_letter = get_column_letter(i)
    sh[f"{col_letter}{max_row+1}"].value = (
        f"=SUM({col_letter}{min_row+1}:{col_letter}{max_row})"
    )

# Apply number format to all cells in the data range, excluding header row (assumed to be the first row)
for row in sh.iter_rows(min_row, max_row + 1, min_column, max_column):
    for cell in row:
        cell.number_format = "#,##0"

wb.save("reports/report.xlsx")
