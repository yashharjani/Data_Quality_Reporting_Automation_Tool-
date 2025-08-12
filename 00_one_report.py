import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from pandas import ExcelWriter
from datetime import datetime
import os
import sys

# Determine the application path
if getattr(sys, "frozen", False):
    application_path = os.path.dirname(sys.executable)
else:
    application_path = os.path.dirname(os.path.abspath(__file__))

# Define the paths to the database and reports folders
db_path = os.path.join(application_path, "db", "Tablero DB_240210.xlsx")
reports_path = os.path.join(application_path, "reports")

# Ensure the reports directory exists
os.makedirs(reports_path, exist_ok=True)

# Load the dataframe
df = pd.read_excel(db_path)
df = df[["Sigla Esp", "Tipo HH", "Horas"]]

# Pivot table creation
pivot_table = df.pivot_table(
    index="Tipo HH", columns="Sigla Esp", values="Horas", aggfunc="sum"
)

# Saving the pivot table to an Excel file
pivot_table_path = os.path.join(reports_path, "pivot_table.xlsx")
with ExcelWriter(pivot_table_path, engine="openpyxl") as writer:
    pivot_table.to_excel(writer, sheet_name="Report", startrow=4)

# Load the workbook to modify
wb = load_workbook(pivot_table_path)
sh = wb["Report"]

# Use direct references from `sh`
min_column = sh.min_column
max_column = sh.max_column
min_row = sh.min_row
max_row = sh.max_row

# BarChart creation
barchart = BarChart()
barchart.title = "Horas por especialidad"
barchart.style = 2

data = Reference(
    sh, min_col=min_column + 1, max_col=max_column, min_row=min_row, max_row=max_row
)
categories = Reference(
    sh, min_col=min_column, max_col=min_column, min_row=min_row + 1, max_row=max_row
)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

sh.add_chart(barchart, "A10")

# Adding sum formulas and applying number format
for i in range(min_column + 1, max_column + 1):
    col_letter = get_column_letter(i)
    sh[f"{col_letter}{max_row+1}"].value = (
        f"=SUM({col_letter}{min_row+1}:{col_letter}{max_row})"
    )

for row in sh.iter_rows(
    min_row=min_row + 1, max_row=max_row + 1, min_col=min_column, max_col=max_column
):
    for cell in row:
        cell.number_format = "#,##0"

# Header setup
sh["A1"] = "Reporte de Horas de IG"
sh["A2"] = "Horas Internas y Subcontratadas"
sh["A1"].font = Font(name="Segoe UI", bold=True, size=12)
sh["A2"].font = Font(name="Segoe UI", bold=True, size=8)

# Get the current date
now = datetime.now()

# Format the date as yymmdd
periodo = now.strftime("%y%m%d")  # Example: '230210'

# Define the output path using the reports_path
output_path = os.path.join(reports_path, f"report_{periodo}.xlsx")

wb.save(output_path)

# application_path = os.path.dirname(sys.executable)

# periodo = input('Enter the periodo')

## Pivot table creation
# pivot_table = df.pivot_table(
#    index="Tipo HH", columns="Sigla Esp", values="Horas", aggfunc="sum"
# )
#
## Saving the pivot table to an Excel file
# with ExcelWriter("reports/pivot_table.xlsx", engine="openpyxl") as writer:
#    pivot_table.to_excel(writer, sheet_name="Report", startrow=4)
#
# input_path = os.path.join(application_path)
#
# wb = load_workbook(input_path)
# sh = wb["Report"]
