from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

wb = load_workbook("reports/pivot_table.xlsx")
sh = wb["Report"]

# Use direct references from `sh` for min and max rows/columns
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

barchart = BarChart()

data = Reference(
    sh, min_col=min_column + 1, max_col=max_column, min_row=min_row, max_row=max_row
)

categories = Reference(
    sh, min_col=min_column, max_col=min_column, min_row=min_row + 1, max_row=max_row
)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

sh.add_chart(barchart, "A10")

barchart.title = "Horas por especialidad"
barchart.style = 2
wb.save("reports/barchart.xlsx")
